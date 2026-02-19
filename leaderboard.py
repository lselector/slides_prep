#!/usr/bin/env python3
"""Extract top 25 from Arena AI leaderboards.

Usage:
  python leaderboard.py

Description:
  Fetches the Arena AI leaderboard pages
  using curl, extracts the top 25 models
  with their scores, and saves results
  into Excel files (.xlsx).

  Each model name in the Excel output is
  a clickable hyperlink to its Arena page.

Output files:
  leaderboard_text.xlsx
  leaderboard_text_coding.xlsx

Requirements:
  pip install openpyxl
"""

import re
import subprocess
import sys
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment
)

URLS = {
    "https://arena.ai/leaderboard/text": (
        "leaderboard_text.xlsx"
    ),
    "https://arena.ai/leaderboard/text/coding": (
        "leaderboard_text_coding.xlsx"
    ),
}

NUM_ROWS = 25
FONT_NAME = "Calibri"
FONT_SIZE = 8

VENDOR_RULES = [
    ("anthropic", ["claude"]),
    ("google", [
        "gemini", "gemma", "palm", "bard"
    ]),
    ("openai", [
        "gpt-", "chatgpt", "o1-", "o3-"
    ]),
    ("opensource", [
        "llama", "mistral", "qwen",
        "deepseek", "yi-", "phi-",
        "command-r", "dbrx", "falcon",
        "vicuna", "wizardlm", "solar",
        "mixtral", "olmo", "jamba",
        "internlm", "glm", "kimi",
    ]),
]

VENDOR_COLORS = {
    "anthropic": "4472C4",
    "google":    "FF0000",
    "openai":    "FFD700",
    "opensource": "00B050",
    "other":     "D9D9D9",
}

# --------------------------------------------------------------
def fetch_html(url):
    """Fetch HTML content from a URL."""
    result = subprocess.run(
        ["curl", "-s", url],
        capture_output=True,
        text=True,
        timeout=30,
    )
    if result.returncode != 0:
        print(f"Error fetching {url}")
        sys.exit(1)
    return result.stdout

# --------------------------------------------------------------
def extract_models(html):
    """Extract model names and URLs from HTML."""
    pattern = (
        r'href="([^"]+)"[^>]*'
        r'title="([^"]+)">'
        r'<span class="truncate max-w-full">'
    )
    matches = re.findall(pattern, html)
    names = [m[1] for m in matches]
    urls = [m[0] for m in matches]
    return names, urls

# --------------------------------------------------------------
def extract_scores(html):
    """Extract score values from the HTML."""
    pattern = (
        r'<span class="text-sm">(\d{4})</span>'
        r'<span class="text-tertiary text-xs">'
        r'(±\d+)</span>'
    )
    return re.findall(pattern, html)

# --------------------------------------------------------------
def get_vendor(name):
    """Determine vendor category from name."""
    lower = name.lower()
    for vendor, keywords in VENDOR_RULES:
        for kw in keywords:
            if kw in lower:
                return vendor
    return "other"

# --------------------------------------------------------------
def make_color_fill(vendor):
    """Create a PatternFill for a vendor."""
    color = VENDOR_COLORS.get(
        vendor, VENDOR_COLORS["other"]
    )
    return PatternFill(
        start_color=color,
        end_color=color,
        fill_type="solid",
    )

# --------------------------------------------------------------
def save_to_excel(names, urls, scores,
                  num_rows, output_file):
    """Save leaderboard data to Excel file."""
    count = min(
        len(names), len(urls),
        len(scores), num_rows
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Leaderboard"
    base_font = Font(
        name=FONT_NAME, size=FONT_SIZE
    )
    header_font = Font(
        name=FONT_NAME, size=FONT_SIZE,
        bold=True
    )
    link_font = Font(
        name=FONT_NAME, size=FONT_SIZE,
        color="0563C1", underline="single"
    )
    ws["A1"] = "Code"
    ws["B1"] = "Model"
    ws["C1"] = "Score"
    for c in [ws["A1"], ws["B1"], ws["C1"]]:
        c.font = header_font
    for i in range(count):
        row = i + 2
        name = names[i]
        score_val = scores[i][0]
        url = urls[i]
        vendor = get_vendor(name)
        cell_a = ws.cell(
            row=row, column=1, value="■"
        )
        cell_a.font = Font(
            name=FONT_NAME, size=FONT_SIZE,
            color=VENDOR_COLORS.get(
                vendor,
                VENDOR_COLORS["other"]
            ),
        )
        cell_a.fill = make_color_fill(vendor)
        cell_b = ws.cell(
            row=row, column=2, value=name
        )
        cell_b.hyperlink = url
        cell_b.font = link_font
        ws.cell(
            row=row, column=3,
            value=int(score_val)
        ).font = base_font
    no_pad = Alignment(
        indent=0, wrap_text=False,
        vertical="center"
    )
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 6
    for r in range(1, count + 2):
        ws.row_dimensions[r].height = 11
    for row_cells in ws.iter_rows(
        min_row=1, max_row=count + 1,
        min_col=1, max_col=3
    ):
        for cell in row_cells:
            cell.alignment = no_pad
    wb.save(output_file)

# --------------------------------------------------------------
def process_url(url, output_file):
    """Fetch, extract, and save data."""
    print(f"Fetching {url} ...")
    html = fetch_html(url)
    names, model_urls = extract_models(html)
    scores = extract_scores(html)
    print(
        f"  Found {len(names)} models, "
        f"{len(scores)} scores"
    )
    save_to_excel(
        names, model_urls, scores,
        NUM_ROWS, output_file
    )
    print(f"  Saved to {output_file}")

# --------------------------------------------------------------
def main():
    """Main function."""
    for url, output_file in URLS.items():
        process_url(url, output_file)
    print("\nDone!")

# --------------------------------------------------------------
if __name__ == "__main__":
    main()
